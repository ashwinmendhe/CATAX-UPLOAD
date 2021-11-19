from django.shortcuts import render
from .models import CataxDB
from .resources import CataxDBResource
from django.contrib import messages
from tablib import Dataset
import openpyxl
from openpyxl import Workbook
import datetime
# Create your views here.
i = 1
j = 1
#-------------main function---------------------------
def main(shRead1,row,column):
    global i, j
    for i in range(1, row+1):
        for j in range(1,column+1):
            txnExchangeDateTime =shRead1.cell(i,1).value
            python_datetime='{:%Y-%m-%d %H:%M:%S}'.format(datetime.datetime.now())
    
            accountID='koinexID'
            accountType = 'Exchange'
            userID='userID'
            txnEntryRoute='Import'
            feeCurrency = 'INR'
            txnVersion = '1.0'
            txnStatus = 'Processing'
            ExchangeName='Koinex Exchange'
            WhosWallet='Customer Crypto wallet'
            valuenew = shRead1.cell(i,j).value
            txnExchangeMemo = 'Success'
            totalValueofTransaction = shRead1.cell(i,6).value
            
            CurrencyName = whichcoin(shRead1.cell(i,2).value)
            exchangeTxnID = 'Null'
            
#---------for credit-------------------------------------------------
            if valuenew == 'BUY' or valuenew == 'Recieve' or valuenew == 'Welcome':
                txnType = 'Credit'
                txnSubType=credit(valuenew)
                creditedCoins =  shRead1.cell(i,4).value
                debitBaseAmount = shRead1.cell(i,5).value
                
                val=CataxDB(accountID=accountID,accountType = accountType,
                    userID=userID, txnEntryRoute=txnEntryRoute,txnType = txnType,txnSubType= txnSubType,
                    creditedCoins=creditedCoins,txnExchangeDate =txnExchangeDateTime,
                    debitBaseAmount=debitBaseAmount,txnExchangeMemo =txnExchangeMemo,
                    creditCurrency=CurrencyName,debitedFromAccountID=ExchangeName,feeCurrency = feeCurrency,
                    totalValueofTransaction = totalValueofTransaction,txnStatus = txnStatus,txnVersion =txnVersion,
                    toCryptoWallet=WhosWallet, createdOn = python_datetime,exchangeTxnID = exchangeTxnID, 
                    txnHash=exchangeTxnID)
                val.save()  

#---------for Debit-------------------------------------------------                
            else:
                if valuenew == 'SELL' or valuenew == 'Send' or valuenew =='Sell Cancel':
                    txnType = 'Debit'
                    txnSubType=debit(valuenew) 
                    debitCoins = shRead1.cell(i,4).value
                    creditBaseAmount = shRead1.cell(i,5).value                
                    
                    val=CataxDB(accountID=accountID,accountType = accountType,
                    userID=userID, txnEntryRoute=txnEntryRoute,txnType = txnType,txnSubType= txnSubType,
                        txnExchangeDate =txnExchangeDateTime,txnExchangeMemo =txnExchangeMemo,
                        debitCoins=debitCoins,creditBaseAmount=creditBaseAmount,
                        debitCurrency=CurrencyName,creditedToAccountID=ExchangeName,feeCurrency = feeCurrency,
                        totalValueofTransaction = totalValueofTransaction,txnStatus = txnStatus,txnVersion =txnVersion,
                        fromCryptoWallet=WhosWallet, createdOn = python_datetime,exchangeTxnID = exchangeTxnID, 
                        txnHash=exchangeTxnID)
                    val.save()

#-------------End---------------------------------------


#-------------credit function---------------------------
def credit(typecredit):
    if typecredit == 'BUY':
           txnSubType = 'BUY'
           return txnSubType
    else:
        if typecredit == 'Recieve':
            txnSubType = 'Deposite'
            return txnSubType
            #shWrite1['G{}'.format(i)].value = 'Deposite'        
        else: 
            if typecredit == 'Welcome':
                txnSubType = 'Reward'
                return txnSubType
           
#-------------End---------------------------------------


#-------------Debit function---------------------------
def debit(typedebit):
    if typedebit == 'Send':
        txnSubType =  'Transfer'
        return txnSubType    

    else:
        if typedebit == 'SELL':
            txnSubType =  'Sell'
            return txnSubType   
        
    
#-------------End---------------------------------------


#-------------Found coin name---------------------------  
def whichcoin(findcoin):
    if findcoin == 'AE/INR':
        newcoin = 'AE'
        return newcoin
    else:
        if findcoin == 'AION/INR':
            newcoin = 'AION'
            return newcoin
        else:
            if findcoin == 'BAT/INR':
                newcoin = 'BAT'
                return newcoin
            else:
                if findcoin == 'BCH/INR':
                    newcoin = 'BCH'
                    return newcoin 
                else:
                    if findcoin == 'BTC/INR':
                     newcoin = 'BTC'
                     return newcoin
                    else:
                        if findcoin == 'EOS/INR':
                         newcoin = 'EOS'
                         return newcoin
                        else:
                            if findcoin == 'ETH/INR':
                             newcoin = 'ETH'
                             return newcoin
                            else:
                                if findcoin == 'GNT/INR':
                                 newcoin = 'GNT'
                                 return newcoin
                                else:
                                    if findcoin == 'LTC/INR':
                                     newcoin = 'LTC'
                                     return newcoin
                                    else:
                                        if findcoin == 'NCASH/INR':
                                         newcoin = 'NCASH'
                                         return newcoin
                                        else:
                                            if findcoin == 'NEO/INR':
                                             newcoin = 'NEO'
                                             return newcoin
                                            else:
                                                if findcoin == 'OMG/INR':
                                                 newcoin = 'OMG'
                                                 return newcoin
                                                else:
                                                    if findcoin == 'REQ/INR':
                                                     newcoin = 'REQ'
                                                     return newcoin
                                                    else:
                                                        if findcoin == 'TRX/INR':
                                                         newcoin = 'TRX'
                                                         return newcoin
                                                        else:
                                                            if findcoin == 'XLM/INR':
                                                             newcoin = 'XLM'
                                                             return newcoin
                                                            else:
                                                                if findcoin == 'XRB/INR':
                                                                 newcoin = 'XRB'
                                                                 return newcoin
                                                                else:
                                                                    if findcoin == 'XRP/INR':
                                                                     newcoin = 'XRP'
                                                                     return newcoin
                                                                    else:
                                                                        if findcoin == 'ZRX/INR':
                                                                         newcoin = 'ZRX'
                                                                         return newcoin
                           
                           

#-------------End---------------------------------------
  
#-------------Data Read--------------------------- 
def dataRead(new_data):
    wbRead= openpyxl.load_workbook(new_data)
    sheets=wbRead.sheetnames
    shRead1= wbRead[sheets[0]]
    row = shRead1.max_row
    column = shRead1.max_column
    return shRead1,row,column
#-------------End---------------------------------------


#-------------Main Coding Start here---------------------------  
def simple_upload(request):
    global wbWrite
    if request.method == "POST":
        new_data = request.FILES['myfiles']
        shRead1,row,column=dataRead(new_data)
        main(shRead1,row,column)
    return render(request, 'upload.html')


#-------------End---------------------------------------