import openpyxl, csv
from pyexcel.cookbook import merge_all_to_a_book
import glob
import pandas as pd
import numpy as np
import datetime
from datetime import datetime, timedelta
x='ADAAaaBTC'

#merge_all_to_a_book(glob.glob("bitbnsBuySale.csv"), "output.xlsx")
#print(x[-4:])
# if 'INR' or 'WRX' or 'BTC' or 'P2P' or 'STF' in x[-3:]:
#     print(x[0:-3],x[-3:] )
# else:
#     print(x[0:-4],x[-4:] )
    

# if 'USDT' in x[-4:]:
#     print(x[0:-4],x[-4:])
# else:
#     print(x[0:-3],x[-3:])



# Reading the csv file
file = ('bitbnsBuySale.csv')
f2 = 'AION/INR'

#read_file = pd.read_csv (r'Path where the CSV file is stored\File name.csv')



df_new = pd.read_csv(r'bitbnsBuySale.csv')
df_new.to_excel (r'Names2.xlsx', index = None, header=True)
# if '.' in file:
#     #x1 =file.index('.')
#     x1 = file[:file.index('.')]
#     #print(file[:file.index('.')])
# print(x1) 
if '/' in f2:
    x2 = f2[:f2.index('/')]
    x3 = f2[f2.index('/')+1:]
    #print(x2, x3)  
#print(len(file))
# saving xlsx file
GFG = pd.ExcelWriter('Names.xlsx')
df_new.to_excel(GFG, index = False)

GFG.save()
wbRead= openpyxl.load_workbook('Names.xlsx')
# wb = openpyxl.Workbook()
# ws = wb.active
for_sheets=wbRead.sheetnames
sheets =for_sheets[0]
#print("sheet name : ", sheets, for_sheets)
# with open('bitbnsBuySale.csv') as f:
#     reader = csv.reader(f, delimiter=':')
#     for row in reader:
#         ws.append(row)

# wb.save('bitbnsBuySale.xlsx')
time1 = '2021-11-26 00:32:18 UTC'
gmt1 = datetime.strptime(time1, '%Y-%m-%d %H:%M:%S UTC')
DateTime1 =gmt1.strftime('%Y-%m-%d %H:%M:%S %Z')

time = 'Fri Sep 04 04:43:27 UTC 2020'
gmt = datetime.strptime(time, '%a %b %d %H:%M:%S UTC %Y')
DateTime =gmt.strftime('%Y-%m-%d %H:%M:%S %Z')
print(DateTime)

time2 = '2021-12-01T07:37:15.766Z'
gmt2 = datetime.strptime(time2, '%Y-%m-%dT%H:%M:%S.%f%z')
DateTime2 =gmt2.strftime('%Y-%m-%d %H:%M:%S')
print(DateTime2)
