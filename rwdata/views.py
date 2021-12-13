from django.shortcuts import render,redirect
from django.http.response import HttpResponseRedirect
from .forms import CreateUserForm
from django.contrib.auth import authenticate, login, logout
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required

from .models import CataxDBnew
from .resources import CataxDBnewResource
from django.contrib import messages
from tablib import Dataset
import openpyxl, csv
from openpyxl import Workbook
import datetime
from datetime import datetime, timedelta
from email.utils import parsedate_tz, mktime_tz
import pandas as pd
import numpy as np
# Create your views here.
i = 1
j = 1
newrow=0

#-----------Login, register, Logout------------------

def registerPage(request):
    if request.user.is_authenticated:
        return redirect('/simple_upload')
        #return render(request,'upload.html')
    else:
        form = CreateUserForm()  
        if request.method == 'POST':
            form = CreateUserForm(request.POST)
            
            if form.is_valid():
                form.save()
            
                user = form.cleaned_data.get('first_name')
                username = form.cleaned_data.get('username')
                messages.success(request,'Account was created for ' + user + ' and his username is: '+ username)
            
                return redirect('login')
        context = {'form':form}
        return render(request, 'register.html',context)

def loginPage(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect(redirect_to="/simple_upload")
    else:
        if request.method == 'POST':
            username=request.POST.get('username')
            password=request.POST.get('password')

            user1 = authenticate(request, username=username, password=password)
            if user1 is not None:
                login(request, user1)
                return HttpResponseRedirect(redirect_to="/simple_upload")
            else:
                messages.info(request, 'Username or password incorrect')
                #return render(request, 'accounts/login.html')
        return render(request, 'login.html')


def logoutUser(request):
    logout(request)
    return redirect('login')


#--------------End-----------------------------------


def buyucoinmain(shRead1,row,column,sheets,user_id,account_id):
    global i, j
    for i in range(2, row+1):
        time2 = shRead1.cell(i,6).value
        gmt2 = datetime.strptime(time2, '%Y-%m-%dT%H:%M:%S.%f%z')
        txnExchangeDateTime =gmt2.strftime('%Y-%m-%d %H:%M:%S')   
        accountID=account_id
        accountType = 'Exchange'
        userID=user_id
        txnEntryRoute='Import'
        sheet_name = sheets
        txnVersion = '1'
        txnStatus = 'Processing'
            
        WhosWallet='Customer Crypto wallet'
        WhoAccountID = 'Customer Crypto ID'
        valuenew = shRead1.cell(i,5).value
        exchangeTxnID = shRead1.cell(i,1).value
        
        feeCurrency = 'Crypto'
        
#-----------------for buy and sell--------------------------------------------------------

        if valuenew == 'BUY' and sheets == 'BuyUCoinBuySell.xlsx':
            txnSubType = 'Buy'
            txnType = 'Credit'
            creditedCoins =  shRead1.cell(i,8).value
            debitBaseAmount = shRead1.cell(i,7).value
            txnExchangeMemo = 'Success: ' +  shRead1.cell(i,3).value
            f2 = shRead1.cell(i,2).value
            totalValueCurrency = f2[:f2.index('-')]
            CurrencyName = f2[f2.index('-')+1:]
            totalValueofTransaction = shRead1.cell(i,7).value * shRead1.cell(i,8).value
            
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=txnSubType,txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,creditedCoins=creditedCoins,txnExchangeDate =txnExchangeDateTime,
            debitBaseAmount=debitBaseAmount,txnExchangeMemo =txnExchangeMemo,creditCurrency=CurrencyName,
            feeCurrency = CurrencyName,totalValueofTransaction = totalValueofTransaction,
            txnStatus = txnStatus,txnVersion =txnVersion,toCryptoWallet=WhosWallet,
            creditedToAccountID=WhoAccountID,exchangeTxnID = exchangeTxnID,txnHash='Null',
            totalValueCurrency = totalValueCurrency,debitBaseCurrency=CurrencyName)
            val.save()  
        elif valuenew == 'SELL' and sheets == 'BuyUCoinBuySell.xlsx':
            txnType = 'Debit'
            txnSubType='Sell'
            debitCoins = shRead1.cell(i,8).value
            creditBaseAmount = shRead1.cell(i,7).value 
            txnExchangeMemo = 'Success: ' +  shRead1.cell(i,3).value
            f2 = shRead1.cell(i,2).value
            totalValueCurrency = f2[:f2.index('-')]
            CurrencyName = f2[f2.index('-')+1:]
        
            totalValueofTransaction = shRead1.cell(i,7).value * shRead1.cell(i,8).value
            
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=valuenew, txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,txnExchangeDate =txnExchangeDateTime,txnExchangeMemo =txnExchangeMemo,
            debitCoins=debitCoins,creditBaseAmount=creditBaseAmount,debitCurrency=CurrencyName,
            debitedFromAccountID=WhoAccountID,feeCurrency = CurrencyName,
            totalValueofTransaction = totalValueofTransaction,txnStatus = txnStatus,txnVersion =txnVersion,
            fromCryptoWallet=WhosWallet,exchangeTxnID = exchangeTxnID,txnHash='Null',
            totalValueCurrency = totalValueCurrency,creditBaseCurrency=CurrencyName)
            val.save()
#-----------------for withdraw and deposit--------------------------------------------------------
           
        elif sheets == 'BuyUCoinWithdraw.xlsx':
            txnType = 'Debit'
            txnSubType='Withdraw'
            debitCoins = shRead1.cell(i,8).value
            txnExchangeMemo = 'Success: ' +  str(shRead1.cell(i,2).value)
            CurrencyName =shRead1.cell(i,3).value
            fee = shRead1.cell(i,9).value
            txnhash = shRead1.cell(i,7).value
            
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=valuenew, txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,txnExchangeDate =txnExchangeDateTime,txnExchangeMemo =txnExchangeMemo,
            debitCoins=debitCoins,debitCurrency=CurrencyName,fees =fee,
            debitedFromAccountID=WhoAccountID,feeCurrency = CurrencyName,
            txnStatus = txnStatus,txnVersion =txnVersion,
            fromCryptoWallet=WhosWallet,exchangeTxnID = exchangeTxnID,txnHash=txnhash)
            val.save()
            
        elif sheets == 'BuyUCoinDeposit.xlsx':
            txnType = 'Credit'
            txnSubType='Deposit'
            creditedCoins = shRead1.cell(i,8).value
            txnExchangeMemo = 'Success: ' +  str(shRead1.cell(i,2).value)
            CurrencyName =shRead1.cell(i,3).value
            fee = shRead1.cell(i,9).value
            txnhash = shRead1.cell(i,7).value
            
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=valuenew, txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,txnExchangeDate =txnExchangeDateTime,txnExchangeMemo =txnExchangeMemo,
            creditedCoins=creditedCoins,creditCurrency=CurrencyName,fees =fee,
            debitedFromAccountID=WhoAccountID,feeCurrency = CurrencyName,
            txnStatus = txnStatus,txnVersion =txnVersion,
            fromCryptoWallet=WhosWallet,exchangeTxnID = exchangeTxnID,txnHash=txnhash)
            val.save()

def coindcxmain(shRead1,row,column,sheets,user_id,account_id):
    global i
    for i in range(2, row+1):

        #txnExchangeDateTime = shRead1.cell(i,7).value
        #print("in dcx main : ", txnExchangeDateTime)
        accountID=account_id
        accountType = 'Exchange'
        userID=user_id
        txnEntryRoute='Import'
        txnVersion = '1'
        txnStatus = 'Processing'
            
        WhosWallet='Customer Crypto wallet'
        WhoAccountID = 'Customer Crypto ID'
        valuenew = shRead1.cell(i,2).value
        
        CurrencyName =shRead1.cell(i,1).value
        totalValueCurrency='INR'
        

#-----------------for buy,sell,Deposit,withdraw --------------------------------------------------------

        if valuenew == 'buy' and sheets == 'Insta_history.xlsx':
            txnType = 'Credit'
            txnSubType = 'Buy'
            #txnSubType = bitbnsDW(sheets)
            creditedCoins =  shRead1.cell(i,3).value
            debitBaseAmount = shRead1.cell(i,4).value
            txnExchangeMemo = shRead1.cell(i,6).value
            exchangeTxnID = 'Null'
            totalValueofTransaction = shRead1.cell(i,5).value
            
            gmt_date = shRead1.cell(i,8).value
            gmt = datetime.strptime(gmt_date, '%Y-%m-%d %H:%M:%S UTC')
            txnExchangeDateTime =gmt.strftime('%Y-%m-%d %H:%M:%S %Z')
            
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=txnSubType,txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,creditedCoins=creditedCoins,txnExchangeDate =txnExchangeDateTime,
            debitBaseAmount=debitBaseAmount,txnExchangeMemo =txnExchangeMemo,creditCurrency=CurrencyName,
            feeCurrency = CurrencyName,totalValueofTransaction = totalValueofTransaction,
            txnStatus = txnStatus,txnVersion =txnVersion,toCryptoWallet=WhosWallet,
            creditedToAccountID=WhoAccountID,exchangeTxnID = exchangeTxnID,txnHash=exchangeTxnID,
            totalValueCurrency = totalValueCurrency, debitBaseCurrency=CurrencyName)
            val.save()  
            
        elif valuenew == 'sell' and sheets == 'Insta_history.xlsx':
            txnType = 'Debit'
            txnSubType='Sell'
            debitCoins = shRead1.cell(i,3).value
            creditBaseAmount = shRead1.cell(i,4).value
            txnExchangeMemo = shRead1.cell(i,6).value
            exchangeTxnID = 'Null'
            totalValueofTransaction = shRead1.cell(i,5).value
            
            gmt_date = shRead1.cell(i,8).value
            gmt = datetime.strptime(gmt_date, '%Y-%m-%d %H:%M:%S UTC')
            txnExchangeDateTime =gmt.strftime('%Y-%m-%d %H:%M:%S %Z')
            
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=valuenew, txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,txnExchangeDate =txnExchangeDateTime,txnExchangeMemo =txnExchangeMemo,
            debitCoins=debitCoins,creditBaseAmount=creditBaseAmount,debitCurrency=CurrencyName,
            debitedFromAccountID=WhoAccountID,feeCurrency = CurrencyName,
            totalValueofTransaction = totalValueofTransaction,txnStatus = txnStatus,txnVersion =txnVersion,
            fromCryptoWallet=WhosWallet,exchangeTxnID = exchangeTxnID,txnHash=exchangeTxnID,
            totalValueCurrency = totalValueCurrency,creditBaseCurrency=CurrencyName)
            val.save()
        elif sheets == 'Deposit_history.xlsx':
            gmt_date = shRead1.cell(i,8).value
            gmt = datetime.strptime(gmt_date, '%Y-%m-%d %H:%M:%S UTC')
            txnExchangeDateTime =gmt.strftime('%Y-%m-%d %H:%M:%S %Z')
            
            txnType = 'Credit'
            txnSubType='Deposit' 
            creditedCoins =  shRead1.cell(i,3).value
            
            txnExchangeMemo = shRead1.cell(i,5).value 
            exchangeTxnID = shRead1.cell(i,4).value 
            debitedFromAccountID,fromCryptoWallet=wallet(shRead1.cell(i,7).value) 
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=txnSubType,txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,creditedCoins=creditedCoins,txnExchangeDate =txnExchangeDateTime,
            txnExchangeMemo =txnExchangeMemo,creditCurrency=CurrencyName,
            feeCurrency = CurrencyName,fromCryptoWallet=fromCryptoWallet, debitedFromAccountID=debitedFromAccountID,
            txnStatus = txnStatus,txnVersion =txnVersion,toCryptoWallet=WhosWallet,
            creditedToAccountID=WhoAccountID,exchangeTxnID = exchangeTxnID,txnHash='Null',
            totalValueCurrency = totalValueCurrency)
            val.save()
        elif sheets == 'Withdraw_history.xlsx':
            gmt_date1 = shRead1.cell(i,10).value
            gmt1 = datetime.strptime(gmt_date1, '%Y-%m-%d %H:%M:%S UTC')
            txnExchangeTime =gmt1.strftime('%Y-%m-%d %H:%M:%S %Z')  
            
            txnType = 'Debit'
            txnSubType='Withdraw'
            debitCoins =  valuenew
            
            txnExchangeMemo = shRead1.cell(i,7).value 
            exchangeTxnID = shRead1.cell(i,4).value 
            AccountID,CryptoWallet=wallet(shRead1.cell(i,9).value) 
            fees  = shRead1.cell(i,3).value
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=txnSubType,txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,debitCoins=debitCoins,txnExchangeDate =txnExchangeTime,
            txnExchangeMemo =txnExchangeMemo,creditCurrency=CurrencyName, fees = fees,
            feeCurrency = CurrencyName,fromCryptoWallet=WhosWallet, debitedFromAccountID=WhoAccountID,
            txnStatus = txnStatus,txnVersion =txnVersion,toCryptoWallet=CryptoWallet,
            creditedToAccountID=AccountID,exchangeTxnID = exchangeTxnID,txnHash='Null',
            totalValueCurrency = totalValueCurrency)
            val.save()

def bitbnsmain(shRead1,row,column,sheets,user_id,account_id):
    global i, j
    for i in range(2, row+1):
        txnExchangeDateTime = shRead1.cell(i,1).value
        accountID=account_id
        accountType = 'Exchange'
        userID=user_id
        txnEntryRoute='Import'
        sheet_name = sheets
        txnVersion = '1'
        txnStatus = 'Processing'
            
        WhosWallet='Customer Crypto wallet'
        WhoAccountID = 'Customer Crypto ID'
        valuenew = shRead1.cell(i,11).value
        reference_no = shRead1.cell(i,8).value
        txnExchangeMemo = 'Success : ' + str(reference_no)
        feeCurrency = shRead1.cell(i,10).value
        CurrencyName =shRead1.cell(i,9).value
        totalValueCurrency=shRead1.cell(i,10).value
        exchangeTxnID = 'Null'
        totalValueofTransaction = shRead1.cell(i,4).value
        fee = shRead1.cell(i,6).value
#-----------------for buy and sell--------------------------------------------------------

        if (valuenew == 'BUY' and sheets == 'bitbnsBuySale.xlsx') or (sheets == 'bitbnd deposite Amount.xlsx'):
            txnSubType = bitbnsDW(sheets)
            txnType = 'Credit'
            creditedCoins =  shRead1.cell(i,2).value
            debitBaseAmount = shRead1.cell(i,5).value
            debitBaseCurrency = shRead1.cell(i,9).value  
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=txnSubType,txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,creditedCoins=creditedCoins,txnExchangeDate =txnExchangeDateTime,
            debitBaseAmount=debitBaseAmount,txnExchangeMemo =txnExchangeMemo,creditCurrency=CurrencyName,
            fees=fee,feeCurrency = feeCurrency,totalValueofTransaction = totalValueofTransaction,
            txnStatus = txnStatus,txnVersion =txnVersion,toCryptoWallet=WhosWallet,
            creditedToAccountID=WhoAccountID,exchangeTxnID = exchangeTxnID,txnHash=exchangeTxnID,
            totalValueCurrency = totalValueCurrency,debitBaseCurrency=debitBaseCurrency)
            val.save()  
        elif valuenew == 'SELL' and sheets == 'bitbnsBuySale.xlsx':
            txnType = 'Debit'
            txnSubType='Sell'
            debitCoins = shRead1.cell(i,2).value
            creditBaseAmount = shRead1.cell(i,5).value 
            creditBaseCurrency =shRead1.cell(i,9).value
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=valuenew, txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,txnExchangeDate =txnExchangeDateTime,txnExchangeMemo =txnExchangeMemo,
            debitCoins=debitCoins,creditBaseAmount=creditBaseAmount,debitCurrency=CurrencyName,
            debitedFromAccountID=WhoAccountID,fees=fee,feeCurrency = feeCurrency,
            totalValueofTransaction = totalValueofTransaction,txnStatus = txnStatus,txnVersion =txnVersion,
            fromCryptoWallet=WhosWallet,exchangeTxnID = exchangeTxnID,txnHash=exchangeTxnID,
            totalValueCurrency = totalValueCurrency, creditBaseCurrency=creditBaseCurrency)
            val.save()
            
            
def wazirxmain(shRead1,row,column,sheets,user_id,account_id):
    global i, j
    for i in range(2, row+1):
        txnExchangeDateTime = shRead1.cell(i,1).value
        accountID=account_id
        accountType = 'Exchange'
        userID=user_id
        txnEntryRoute='Import'
        sheet_name = sheets
        txnVersion = '1'
        txnStatus = 'Processing'
            
        WhosWallet='Customer Crypto wallet'
        WhoAccountID = 'Customer Crypto ID'
        valuenew = shRead1.cell(i,6).value
        txnExchangeMemo = 'Success'
#-----------------for buy and sell-------------------------------------------------------
        if sheet_name == 'Exchange Trades' or sheet_name == 'P2P Trades':
            if valuenew == 'Buy':
                txnType = 'Credit'
                txnSubType='Buy'
                creditedCoins =  shRead1.cell(i,4).value
                debitBaseAmount = shRead1.cell(i,3).value
                feeCurrency = shRead1.cell(i,7).value
                exchangeTxnID = 'Null'
                totalValueofTransaction = shRead1.cell(i,5).value
                fee_b = shRead1.cell(i,8).value
                    # write logic from market field
                CurrencyName,totalValueCurrency=crypto1n2(shRead1.cell(i,2).value)

                val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
                txn=txnSubType,txnEntryRoute=txnEntryRoute,txnType = txnType,
                txnSubType= txnSubType,creditedCoins=creditedCoins,txnExchangeDate =txnExchangeDateTime,
                debitBaseAmount=debitBaseAmount,txnExchangeMemo =txnExchangeMemo,creditCurrency=CurrencyName,
                fees=fee_b,feeCurrency = feeCurrency,totalValueofTransaction = totalValueofTransaction,
                txnStatus = txnStatus,txnVersion =txnVersion,toCryptoWallet=WhosWallet,
                creditedToAccountID=WhoAccountID,exchangeTxnID = exchangeTxnID,txnHash=exchangeTxnID,
                totalValueCurrency = totalValueCurrency,debitBaseCurrency=CurrencyName)
                val.save()  
            elif valuenew == 'Sell':
                txnType = 'Debit'
                txnSubType='Sell'
                debitCoins = shRead1.cell(i,4).value
                creditBaseAmount = shRead1.cell(i,3).value 
                feeCurrency = shRead1.cell(i,7).value
                exchangeTxnID = 'Null'
                totalValueofTransaction = shRead1.cell(i,5).value
                fee_s = shRead1.cell(i,8).value

                CurrencyName,totalValueCurrency=crypto1n2(shRead1.cell(i,2).value)

                    
                val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
                txn=txnSubType, txnEntryRoute=txnEntryRoute,txnType = txnType,
                txnSubType= txnSubType,txnExchangeDate =txnExchangeDateTime,txnExchangeMemo =txnExchangeMemo,
                debitCoins=debitCoins,creditBaseAmount=creditBaseAmount,debitCurrency=CurrencyName,
                debitedFromAccountID=WhoAccountID,fees=fee_s,feeCurrency = feeCurrency,
                totalValueofTransaction = totalValueofTransaction,txnStatus = txnStatus,txnVersion =txnVersion,
                fromCryptoWallet=WhosWallet,exchangeTxnID = exchangeTxnID,txnHash=exchangeTxnID,
                totalValueCurrency = totalValueCurrency,creditBaseCurrency=CurrencyName)
                val.save()

#-----------------for Deposits and Withdrawals or Additional Transfers-------------------
            
        elif sheet_name == 'Deposits and Withdrawals' or sheet_name == 'Additional Transfers':
            if shRead1.cell(i,2).value == 'Withdraw':
                txnType = 'Debit'
                txnSubType=shRead1.cell(i,2).value
                debitCoins = shRead1.cell(i,4).value
                exchangeTxnID = shRead1.cell(i,6).value
                txnHash = shRead1.cell(i,8).value
                CurrencyName = shRead1.cell(i,3).value
                feeCurrency=shRead1.cell(i,3).value
                creditedToAccountID= shRead1.cell(i,5).value
                toCryptoWallet =  shRead1.cell(i,7).value

                val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
                txn=txnSubType, txnEntryRoute=txnEntryRoute,txnType = txnType,
                txnSubType= txnSubType,txnExchangeDate =txnExchangeDateTime,txnExchangeMemo =txnExchangeMemo,
                debitCoins=debitCoins,debitCurrency=CurrencyName,creditedToAccountID=creditedToAccountID,
                debitedFromAccountID=WhoAccountID,txnStatus = txnStatus,txnVersion =txnVersion,
                fromCryptoWallet=WhosWallet,toCryptoWallet=toCryptoWallet, exchangeTxnID = exchangeTxnID,
                txnHash=txnHash, feeCurrency=feeCurrency)
                val.save()

            elif shRead1.cell(i,2).value == 'Deposit':
                txnType = 'Credit'
                txSubType=shRead1.cell(i,2).value
                creditCoins = shRead1.cell(i,4).value
                exchangeTxnID = shRead1.cell(i,6).value
                txnHash = shRead1.cell(i,8).value
                CurrencyName = shRead1.cell(i,3).value
                feeCurrency=shRead1.cell(i,3).value

                fromCryptoWallet = shRead1.cell(i,7).value
                debitedFromAccountID,aidrop_remark, subtyp=airdrop(shRead1.cell(i,5).value,sheets)
                val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
                txn=txSubType, txnEntryRoute=txnEntryRoute,txnType = txnType,
                txnSubType= subtyp,txnExchangeDate =txnExchangeDateTime,txnExchangeMemo =txnExchangeMemo,
                creditedCoins=creditCoins,creditCurrency=CurrencyName,creditedToAccountID=WhoAccountID,
                debitedFromAccountID=debitedFromAccountID,txnStatus = txnStatus,txnVersion =txnVersion,
                fromCryptoWallet=fromCryptoWallet,toCryptoWallet=WhosWallet, exchangeTxnID = exchangeTxnID,
                txnHash=txnHash, feeCurrency=feeCurrency, txnCustomerMemo=aidrop_remark)
                val.save()
        elif sheet_name == 'Third party transfers':
                txnType = 'Debit'
                txnSubType='Transfer'
                debitCoins = shRead1.cell(i,3).value
                exchangeTxnID = 'Null'
                txnHash = 'Null'
                CurrencyName = shRead1.cell(i,2).value
                feeCurrency='Null'
                toCryptoWallet =  shRead1.cell(i,5).value
                fromCryptoWallet=shRead1.cell(i,4).value
                
                val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
                txn=txnSubType, txnEntryRoute=txnEntryRoute,txnType = txnType,
                txnSubType= txnSubType,txnExchangeDate =txnExchangeDateTime,txnExchangeMemo =txnExchangeMemo,
                debitCoins=debitCoins,debitCurrency=CurrencyName,creditedToAccountID=WhoAccountID,
                debitedFromAccountID=WhoAccountID,txnStatus = txnStatus,txnVersion =txnVersion,
                fromCryptoWallet=fromCryptoWallet,toCryptoWallet=toCryptoWallet, exchangeTxnID = exchangeTxnID,
                txnHash=txnHash, feeCurrency=feeCurrency)
                val.save()

            
#-------------End--------------------------------------- 

#-------------binance----------------------------------
def binancemain(shRead1,row,column,sheets,user_id,account_id):
    global i, j
    for i in range(2, row+1):
        accountID=account_id
        accountType = 'Exchange'
        userID=user_id
        txnEntryRoute='Import'
        feeCurrency = 'Crypto'
        txnVersion = '1'
        txnStatus = 'Processing'
        txnCustomerMemo = "UserID: " + shRead1.cell(i,1).value
        valuenew = sheets
        txnExchangeMemo = 'Success'
            #totalValueofTransaction = 'Null'
        totalValueCurrency = 'Null'
        CurrencyName = shRead1.cell(i,2).value
        exchangeTxnID = shRead1.cell(i,6).value
            
        toCryptoWallet=shRead1.cell(i,4).value
  #---------for credit-------------------------------------------------
        if valuenew == 'Deposit History':
            txnType = 'Credit'
            txnSubType='Deposit'
            creditedCoins =  shRead1.cell(i,3).value
                #debitBaseAmount = 'Null'              

            C_dID,C_fCWallet = wallet(shRead1.cell(i,5).value) 
            creditedToAccountID = 'Customer binance ID'
                
            txnExchangeDateTime =shRead1.cell(i,7).value
                #python_datetime='{:%Y-%m-%d %H:%M:%S}'.format(datetime.now())
                
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=txnSubType,txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,creditedCoins=creditedCoins,txnExchangeDate =txnExchangeDateTime,
            txnExchangeMemo =txnExchangeMemo,creditCurrency=CurrencyName,debitedFromAccountID=C_dID,
            creditedToAccountID=creditedToAccountID,feeCurrency = feeCurrency,txnStatus = txnStatus,
            txnVersion =txnVersion,toCryptoWallet=toCryptoWallet, fromCryptoWallet=C_fCWallet,
            exchangeTxnID = exchangeTxnID,txnHash=exchangeTxnID,totalValueCurrency = totalValueCurrency,
            txnCustomerMemo=txnCustomerMemo)
            val.save()  

#---------for Debit-------------------------------------------------                
        elif valuenew == 'Withdrawal History':
            txnType = 'Debit'
            txnSubType='Withdrawal'
            debitCoins = shRead1.cell(i,3).value
            #creditBaseAmount = 'Null'
            D_dID = 'Customer binance ID' ; D_fCWallet = 'Customer Crypto wallet'
            creditedToAccountID = 'Null'
                    
            gmt_date = shRead1.cell(i,7).value
                    
                    #print("print withdrwal time",gmt_date, "print i: ",i)
            gmt = datetime.strptime(gmt_date, '%a %b %d %H:%M:%S UTC %Y')

            txnExchangeDateTime =gmt.strftime('%Y-%m-%d %H:%M:%S %Z')
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=txnSubType, txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,txnExchangeDate =txnExchangeDateTime,txnExchangeMemo =txnExchangeMemo,
            debitCoins=debitCoins,debitCurrency=CurrencyName,creditedToAccountID=creditedToAccountID,
            debitedFromAccountID= D_dID,feeCurrency = feeCurrency,txnStatus = txnStatus,
            txnVersion =txnVersion,toCryptoWallet=toCryptoWallet,fromCryptoWallet= D_fCWallet,
            exchangeTxnID = exchangeTxnID,txnHash=exchangeTxnID,totalValueCurrency = totalValueCurrency, 
            txnCustomerMemo=txnCustomerMemo)
            val.save()

#-------------End---------------------------------------
          

#-------------end----------------------------------
#-------------Koinex function---------------------------
def Koinexmain(shRead1,row,column,sheets,user_id,account_id):
    global i, j
    for i in range(2, row+1):
        txnExchangeDateTime =shRead1.cell(i,1).value
        #python_datetime='{:%Y-%m-%d %H:%M:%S}'.format(datetime.now())
        accountID=account_id
        accountType = 'Exchange'
        userID=user_id
        txnEntryRoute='Import'
        feeCurrency = 'Crypto'
        txnVersion = '1'
        txnStatus = 'Processing'
        ExchangeName='Customer koinex ID'
        WhosWallet='Customer Crypto wallet'
        valuenew = shRead1.cell(i,3).value
        txnExchangeMemo = 'Success'
        totalValueofTransaction = shRead1.cell(i,6).value
        #f2 = whichcoin(shRead1.cell(i,2).value)
        f2 = shRead1.cell(i,2).value
        CurrencyName = f2[:f2.index('/')]
        totalValueCurrency = f2[f2.index('/')+1:]
            
#---------for credit-------------------------------------------------
        if valuenew == 'BUY' or valuenew == 'Recieve' or valuenew == 'Welcome':
            txnType = 'Credit'    
            txnSubType=credit(valuenew)
            creditedCoins =  shRead1.cell(i,4).value
            debitBaseAmount = shRead1.cell(i,5).value
                
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID, 
            txn=valuenew,txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,creditedCoins=creditedCoins,txnExchangeDate =txnExchangeDateTime,
            debitBaseAmount=debitBaseAmount,txnExchangeMemo =txnExchangeMemo,creditCurrency=CurrencyName,
            creditedToAccountID=ExchangeName,feeCurrency = feeCurrency,totalValueofTransaction = totalValueofTransaction,
            txnStatus = txnStatus,txnVersion =txnVersion,toCryptoWallet=WhosWallet,
            totalValueCurrency = totalValueCurrency)
            val.save()  

#---------for Debit-------------------------------------------------                
        elif valuenew == 'SELL' or valuenew == 'Send' or valuenew =='Sell Cancel':
            txnType = 'Debit'
            txnSubType=debit(valuenew) 
            debitCoins = shRead1.cell(i,4).value
            creditBaseAmount = shRead1.cell(i,5).value                    
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=valuenew, txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,txnExchangeDate =txnExchangeDateTime,txnExchangeMemo =txnExchangeMemo,
            debitCoins=debitCoins,creditBaseAmount=creditBaseAmount,debitCurrency=CurrencyName,
            debitedFromAccountID=ExchangeName,feeCurrency = feeCurrency,totalValueofTransaction = totalValueofTransaction,
            txnStatus = txnStatus,txnVersion =txnVersion,fromCryptoWallet=WhosWallet,
            totalValueCurrency = totalValueCurrency)
            val.save()

#-------------End---------------------------------------



#-------------Zebpay function---------------------------
def Zebmain(shRead1,row,column,sheet,length,user_id,account_id):
    global i, j
    for i in range(2, row+1):
        txnExchangeDateTime =shRead1.cell(i,1).value           
        accountID=account_id
        accountType = 'Exchange'
        userID=user_id
        txnEntryRoute='Import'
        feeCurrency = 'Crypto'
        txnVersion = '1'
        txnStatus = 'Processing'
        ExchangeName='Customer Zebpay ID'
        WhosWallet='Customer Crypto wallet'
        valuenew = shRead1.cell(i,2).value
        txnExchangeMemo = shRead1.cell(i,5).value
        totalValueCurrency = 'INR'
        CurrencyName = sheet[length]
        exchangeTxnID = shRead1.cell(i,9).value

#---------for credit-------------------------------------------------
        if valuenew == 'Buy' or valuenew == 'Recieve' or valuenew == 'Internal Recieve' or valuenew == 'Welcome' or valuenew =='Sell Cancel':
            txnType = 'Credit'
            txnSubType=credit(valuenew)
            creditedCoins =  shRead1.cell(i,4).value               
            debitBaseAmount = shRead1.cell(i,3).value
            totalValueofTransaction = creditedCoins * debitBaseAmount
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID, 
            txn=valuenew, txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,creditedCoins=creditedCoins,txnExchangeDate =txnExchangeDateTime,
            txnExchangeMemo = txnExchangeMemo,debitBaseAmount =debitBaseAmount,creditCurrency=CurrencyName,
            creditedToAccountID=ExchangeName,feeCurrency = feeCurrency,totalValueofTransaction = totalValueofTransaction,
            txnStatus = txnStatus,txnVersion = txnVersion,toCryptoWallet=WhosWallet,
            exchangeTxnID = exchangeTxnID,totalValueCurrency = totalValueCurrency,debitBaseCurrency=CurrencyName)
            val.save()  

#---------for Debit-------------------------------------------------                
        elif valuenew == 'Sell' or valuenew == 'Send' or valuenew == 'Internal Send':
            txnType = 'Debit'
            txnSubType=debit(valuenew) 
            debitCoins = shRead1.cell(i,4).value              
            creditBaseAmount = shRead1.cell(i,3).value 
            totalValueofTransaction = debitCoins * creditBaseAmount
            val=CataxDBnew(accountID=accountID,accountType = accountType,userID=userID,
            txn=valuenew, txnEntryRoute=txnEntryRoute,txnType = txnType,
            txnSubType= txnSubType,txnExchangeDate =txnExchangeDateTime,txnExchangeMemo = txnExchangeMemo,
            debitCoins=debitCoins,creditBaseAmount=creditBaseAmount,debitCurrency=CurrencyName,
            debitedFromAccountID=ExchangeName,feeCurrency = feeCurrency,totalValueofTransaction = totalValueofTransaction,
            txnStatus = txnStatus,txnVersion =txnVersion,fromCryptoWallet=WhosWallet, 
            exchangeTxnID = exchangeTxnID,totalValueCurrency = totalValueCurrency,creditBaseCurrency=CurrencyName)
            val.save()

#-------------End---------------------------------------

def bitbnsDW(sheets):
    if sheets == 'bitbnd deposite Amount.xlsx':
        return 'Deposit'
    else:
        return 'Buy'

def airdrop(find_airdrop, sheet):
    if sheet == 'Additional Transfers':
        debit = 'Null'
        airdrop = find_airdrop
        if 'Airdrop' or 'airdrop' in find_airdrop:
            return debit,airdrop, 'Airdrop'
        elif 'Reward' in find_airdrop:
            return debit,airdrop, 'Reward'
        elif  'Bonus' in find_airdrop:
            return debit,airdrop, 'Reward'
    else:
        debit =find_airdrop
        airdrop = 'Null'
        return debit,airdrop, 'Deposit'

def wallet(walletvalue):
    if len(str(walletvalue)) == 11 or len(str(walletvalue)) == 7 or len(str(walletvalue)) == 34:                               
        debitedFromAccountID =walletvalue
        fromCryptoWallet = 'Null'
        return debitedFromAccountID,fromCryptoWallet
    else:
        debitedFromAccountID ='Null'
        fromCryptoWallet = walletvalue
        return debitedFromAccountID,fromCryptoWallet   

def crypto1n2(bothcrypto):
    if 'USDT' in bothcrypto[-4:]:
        return bothcrypto[0:-4],bothcrypto[-4:]
    else:
        return bothcrypto[0:-3],bothcrypto[-3:]   
#-------------credit function---------------------------
def credit(typecredit):
    if typecredit == 'BUY' or typecredit == 'Buy':
        return 'Buy'
    elif typecredit == 'Recieve':
        return 'deposit'      
    elif typecredit == 'Welcome':
        return 'Reward'
    elif typecredit == 'Internal Recieve':
        return'deposit'
    elif typecredit == 'Sell Cancel':
        return 'Sell Cancel'
                   
#-------------End---------------------------------------


#-------------Debit function---------------------------
def debit(typedebit):
    if typedebit == 'Send':
        return 'Transfer'  
    elif typedebit == 'SELL' or typedebit == 'Sell':
        return 'Sell'   
    elif typedebit == 'Internal Send':
        return 'Transfer'

#-------------End---------------------------------------

#-------------Data Read--------------------------- 
def dataRead_zebpay(new_data,user_id,account_id):
    global newrow
    wbRead= openpyxl.load_workbook(new_data)  
    sheets=wbRead.sheetnames
    shRead1= [0]*len(sheets)
    for length in range(len(sheets)):
        shRead1[length]= wbRead[sheets[length]]
        row = shRead1[length].max_row
        column = shRead1[length].max_column
        newrow = newrow +row
        Zebmain(shRead1[length],row,column,sheets,length,user_id,account_id)
        print('Succesfully uploaded Zebpay trade data: ', length)
    return newrow
#-------------End---------------------------------------

#-------------Data Read--------------------------- 
def dataRead_koinex(new_data):
    wbRead= openpyxl.load_workbook(new_data)    
    sheets=wbRead.sheetnames
    shRead1= wbRead[sheets[0]]
    row = shRead1.max_row
    column = shRead1.max_column
    return shRead1,row,column,sheets
#-------------End---------------------------------------

#-------------Data Read--------------------------- 
def dataRead_binance(new_data,user_id,account_id):
    global newrow
    wbRead= openpyxl.load_workbook(new_data)    
    sheets=wbRead.sheetnames
    #shRead1= [0]*len(sheets)
    for length in range(len(sheets)):
        if sheets[length] == 'Deposit History' or sheets[length] == 'Withdrawal History':
            shRead1= wbRead[sheets[length]]
            row = shRead1.max_row
            column = shRead1.max_column
            newrow = newrow +row
            binancemain(shRead1,row,column,sheets[length],user_id,account_id)
    return newrow             
#-------------End---------------------------------------

#-------------Data Read--------------------------- 
def dataRead_wzirx(new_data,user_id,account_id):
    global newrow
    wbRead= openpyxl.load_workbook(new_data)    
    sheets=wbRead.sheetnames
    for length in range(0,len(sheets)):
        if sheets[length] == 'Exchange Trades' or sheets[length] == 'P2P Trades' or sheets[length] == 'Deposits and Withdrawals' or sheets[length] == 'Additional Transfers' or  sheets[length] == 'Third party transfers':
            shRead1= wbRead[sheets[length]]
            row = shRead1.max_row
            column = shRead1.max_column
            #print("row: ",row, "column: ",column, "sheet name: ",sheets[length])
            newrow = newrow +row
            wazirxmain(shRead1,row,column,sheets[length],user_id,account_id)
    return newrow   
#-------------End---------------------------------------

#-------------Data Read--------------------------- 
def dataRead_bitbns(new_data,user_id,account_id):
    global newrow
    #wbRead= new_data
    wbRead= openpyxl.load_workbook(new_data)
    for_sheets=wbRead.sheetnames
    sheets =for_sheets[0]
    print("file name : ", new_data)

    shRead1= wbRead[sheets]
    row = shRead1.max_row
    column = shRead1.max_column
    #print("row: ",row, "column: ",column, "sheet name: ",sheets)
    newrow = newrow +row
    bitbnsmain(shRead1,row,column,new_data,user_id,account_id)
    return newrow
#-------------End---------------------------------------

def dataRead_coindcx(new_data,user_id,account_id):
    global newrow
    wbRead= openpyxl.load_workbook(new_data)
    for_sheets=wbRead.sheetnames
    sheets =for_sheets[0]
    print("file name : ", new_data)
    shRead1= wbRead[sheets]
    row = shRead1.max_row
    column = shRead1.max_column
    print("row: ",row, "column: ",column, "sheet name: ",sheets)
    newrow = newrow +row
    coindcxmain(shRead1,row,column,new_data,user_id,account_id)
    return newrow

def dataRead_buyucoin(new_data,user_id,account_id):
    global newrow
    wbRead= openpyxl.load_workbook(new_data)
    for_sheets=wbRead.sheetnames
    sheets =for_sheets[0]
    print("file name : ", new_data)
    shRead1= wbRead[sheets]
    row = shRead1.max_row
    column = shRead1.max_column
    print("row: ",row, "column: ",column, "sheet name: ",sheets)
    newrow = newrow +row
    buyucoinmain(shRead1,row,column,new_data,user_id,account_id)
    return newrow





def convertCSVtoXL(new_dat):
    #print("1..: ",str(new_dat), "2..: ", new_dat)
    if not new_dat.name.endswith('xlsx'):
        df_new = pd.read_csv(new_dat)
        file = str(new_dat)
        filename = "%s.xlsx" % file[:file.index('.')]
        GFG = pd.ExcelWriter(filename)
        df_new.to_excel(GFG, index = False)
        GFG.save()
        #new_dat= openpyxl.load_workbook(filename)
        return filename
    else:
        return new_dat
#-------------Main Coding Start here---------------------------  

#@method_decorator(login_required, name="dispatch") 
def simple_upload(request):
    global newrow
    if request.method == "POST":
        new_dat = request.FILES['myfiles']
        new_data = convertCSVtoXL(new_dat)
        user_id=request.POST.get('id')
        account_id=request.POST.get('Eid')
        exchange_name=request.POST.get('exchange')

        # count=dataRead_coindcx(new_dat,user_id,account_id)
        # return render(request,'result.html',{'res':count}) 
 #--------------for zebpay-----------------------------------       
        #print(new_data,user_id,account_id,exchange_name)
        if str(exchange_name) == 'zebpay':

            count=dataRead_zebpay(new_data,user_id,account_id)
            return render(request,'result.html',{'res':count})
 #--------------zebpay end----------------------------------- 
#--------------for koinex-----------------------------------                       
        elif str(exchange_name)=='koinex':
            shRead1,row,column,sheets=dataRead_koinex(new_data)
            Koinexmain(shRead1,row,column,sheets,user_id,account_id)
            return render(request,'result.html',{'res':row-1})
#--------------koinex end----------------------------------- 
#--------------for binance-----------------------------------          
        elif str(exchange_name)=='binance':
            count=dataRead_binance(new_data,user_id,account_id)
            print(exchange_name)
            return render(request,'result.html',{'res':count-1})
#--------------binance end----------------------------------- 
#--------------for wazirx-----------------------------------                 
        elif str(exchange_name) == 'wazirx':
            count=dataRead_wzirx(new_data,user_id,account_id)
            return render(request,'result.html',{'res':count-1})
#--------------wzirx end------------------------------------
#--------------for bitbns-----------------------------------
        elif str(exchange_name) == 'bitbns':
            count=dataRead_bitbns(new_data,user_id,account_id)
            return render(request,'result.html',{'res':count-1})                    
#--------------bitbns end------------------------------------  
#--------------for coindcx-----------------------------------             
        elif str(exchange_name) == 'coindcx':
            count=dataRead_coindcx(new_data,user_id,account_id)
            return render(request,'result.html',{'res':count-1}) 
#--------------coindcx end------------------------------------  
#--------------for Buyucoin-----------------------------------             
        elif str(exchange_name) == 'buyucoin':
            count=dataRead_buyucoin(new_data,user_id,account_id)
            return render(request,'result.html',{'res':count-1}) 
#--------------buyucoin end------------------------------------         

    return render(request, 'upload.html')
 


#-------------End---------------------------------------